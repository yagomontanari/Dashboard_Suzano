from fastapi import APIRouter, Depends, HTTPException, status, BackgroundTasks, Request
from core.limiter import limiter
from fastapi.security import OAuth2PasswordBearer
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from jose import JWTError, jwt
import logging
import asyncio
import io
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
from openpyxl.styles import PatternFill, Font

from core.database import get_db, AsyncSessionLocal
from core.security import SECRET_KEY, ALGORITHM
from core.mail import mail_service
from core.utils import parse_date_range, apply_excel_premium_style
from api.queries import *
from fastapi.responses import StreamingResponse
from core.config import settings

router = APIRouter()
logger = logging.getLogger(__name__)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

from core.database_app import get_app_db
from core.models_app import User, UserStatus


async def get_current_user(
    token: str = Depends(oauth2_scheme), db_app: AsyncSession = Depends(get_app_db)
):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED, detail="Token invalido"
            )

        # Busca o usuario no banco da aplicacao (Supabase)
        from sqlalchemy.future import select

        result = await db_app.execute(select(User).where(User.email == email))
        user = result.scalar_one_or_none()

        if not user or user.status != UserStatus.APPROVED:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Usuario nao autorizado ou pendente de aprovacao",
            )

        return user
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, detail="Token invalido"
        )


@router.get("/dashboard")
@limiter.limit("60/minute")
async def get_dashboard_metrics(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    db: AsyncSession = Depends(get_db),
    user: str = Depends(get_current_user),
):
    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)

        params_str = {
            "start_date": start_dt.strftime("%Y-%m-%d %H:%M:%S"),
            "end_date": end_dt.strftime("%Y-%m-%d %H:%M:%S"),
        }
        params = {"start_date": start_dt, "end_date": end_dt}

        # Otimizacao: Paralelizando 11 chamadas utilizando conexoes via polling para baixar o tempo de 6s para 1~2s.
        async def fetch_data(q, p, is_count=False):
            async with AsyncSessionLocal() as session:
                if is_count:
                    # Uso Seguro de Bind Parameters para o Wrap da Query
                    # Evitamos f-string aqui para estrutura de SQL onde possivel
                    # O count_query original em api/queries ja e um objeto text()
                    # Mas se precisamos envolver em subquery, fazemos sem injetar strings brutas de input
                    count_wrapper = text(
                        f"SELECT COUNT(1) AS total FROM ({q.text.strip().rstrip(';')}) AS subquery"
                    )
                    res = await session.execute(count_wrapper, p)
                    return res.scalar() or 0
                else:
                    res = await session.execute(q, p)
                    return [dict(row) for row in res.mappings().all()]

        results = await asyncio.gather(
            fetch_data(QUERY_ORCAMENTO_INTEGRACAO_TOTAL, params_str),
            fetch_data(QUERY_ZAJU_TOTAL, params),
            fetch_data(QUERY_PAGAMENTOS_TOTAL, params),
            fetch_data(QUERY_TOP_CLIENTES, params),
            fetch_data(QUERY_DASHBOARD_COUNTS_CONSOLIDATED, params), # 7 contagens em 1 única conexão
            fetch_data(QUERY_LAST_SYNC, {}),
        )

        (
            vk11_res,
            zaju_res,
            zver_res,
            top_clients_res,
            counts_res,
            last_sync_res,
        ) = results

        # Extração das contagens consolidadas
        counts = counts_res[0] if counts_res else {}
        sellin_count = counts.get("sellin", 0)
        clientes_count = counts.get("clientes", 0)
        produtos_count = counts.get("produtos", 0)
        cutoff_count = counts.get("cutoff", 0)
        usuarios_count = counts.get("usuarios", 0)
        pagamentos_count = counts.get("pagamentos", 0)
        vk11_count = counts.get("vk11", 0)

        # Mapeando os resultados
        vk11_totals = (
            vk11_res[0] if vk11_res else {"success": 0, "pending": 0, "error": 0}
        )
        zaju_totals = (
            zaju_res[0]
            if zaju_res
            else {
                "success": 0,
                "pending": 0,
                "error": 0,
                "pending_return": 0,
                "total": 0,
            }
        )
        zver_totals_raw = (
            zver_res[0]
            if zver_res
            else {
                "success": 0,
                "pending": 0,
                "pending_return": 0,
                "error": 0,
                "value_success": 0.0,
                "value_pending": 0.0,
                "value_pending_return": 0.0,
                "value_error": 0.0,
            }
        )

        top_clients = top_clients_res

        zver_totals = {
            "success": zver_totals_raw["success"],
            "pending": zver_totals_raw["pending"],
            "pending_return": zver_totals_raw["pending_return"],
            "error": zver_totals_raw["error"],
            "value_success": float(zver_totals_raw["value_success"]),
            "value_pending": float(zver_totals_raw["value_pending"]),
            "value_pending_return": float(zver_totals_raw["value_pending_return"]),
            "value_error": float(zver_totals_raw["value_error"]),
            "top_clients": top_clients,
        }

        return {
            "source": "postgresql",
            "vk11": dict(vk11_totals),
            "zaju": dict(zaju_totals),
            "zver": zver_totals,
            "errors": {
                "sellin": sellin_count,
                "clientes": clientes_count,
                "produtos": produtos_count,
                "cutoff": cutoff_count,
                "usuarios": usuarios_count,
                "pagamentos": pagamentos_count,
                "vk11": vk11_count,
                "zaju": zaju_totals.get("error", 0),
            },
            "last_updates": [
                {
                    "categoria": {
                        "PRE_CADASTRO_USUARIO": "Usuarios",
                        "CLIENTE": "Clientes",
                        "PAGAMENTO_LIQUIDADO": "Retorno Pagamento",
                        "CUTOFF": "Cutoff",
                        "SELLIN": "Sell-In",
                        "PRODUTO": "Produtos",
                        "ORCAMENTO": "VK11",
                        "PAGAMENTO": "ZVER",
                        "DADOS_PROVISOES": "Dados Provisoes",
                        "AJUSTE_PROVISAO": "ZAJU",
                    }.get(r["categoria"], r["categoria"]),
                    "data": r["dta"].isoformat() if r["dta"] else None,
                    "direcao": r["direcao"],
                    "lote": r["lote"],
                    "mensagem": {
                        "PRE_CADASTRO_USUARIO": f"Sincronizacao de usuarios processada{' (Lote ' + r['lote'] + ')' if r['lote'] else ''}.",
                        "CLIENTE": f"Carga cadastral de clientes recebida{' (Lote ' + r['lote'] + ')' if r['lote'] else ''}.",
                        "PAGAMENTO_LIQUIDADO": f"Retorno de liquidacao SAP recebido{' (Lote ' + r['lote'] + ')' if r['lote'] else ''}.",
                        "CUTOFF": f"Atualizacao de calendario cutoff{' (Lote ' + r['lote'] + ')' if r['lote'] else ''}.",
                        "SELLIN": f"Notas fiscais de Sell-In integradas{' (Lote ' + r['lote'] + ')' if r['lote'] else ''}.",
                        "PRODUTO": f"Dados mestre de produtos atualizados{' (Lote ' + r['lote'] + ')' if r['lote'] else ''}.",
                        "ORCAMENTO": "Envio de orcamentos (VK11) para o SAP.",
                        "PAGAMENTO": "Envio de remessa de pagamentos (ZVER) para o SAP.",
                        "DADOS_PROVISOES": "Sincronizacao de base de provisoes concluida.",
                        "AJUSTE_PROVISAO": "Ajuste de provisao (ZAJU) enviado ao SAP.",
                    }.get(r["categoria"], "Sincronizacao de dados finalizada."),
                }
                for r in last_sync_res
            ],
        }
    except Exception as e:
        logger.error(f"Erro ao conectar ou ler do banco de dados no Dashboard: {e}")
        # Sanitizacao de Resposta: Nao expor traceback ao cliente
        raise HTTPException(
            status_code=500, detail="Erro interno ao processar metricas do dashboard"
        )


@router.get("/dashboard/inconsistencies/{category}")
@limiter.limit("60/minute")
async def get_inconsistencies(
    request: Request,
    category: str,
    start_date: str = None,
    end_date: str = None,
    page: int = 1,
    size: int = 20,
    sort_by: str = None,
    order: str = "desc",
    db: AsyncSession = Depends(get_db),
    user: str = Depends(get_current_user),
):
    try:
        offset = (page - 1) * size
        start_dt, end_dt = parse_date_range(start_date, end_date)

        params = {
            "start_date": start_dt,
            "end_date": end_dt,
            "limit": size,
            "offset": offset,
        }
        params_count = {"start_date": start_dt, "end_date": end_dt}

        # Consistent with get_dashboard_metrics, VK11 needs string dates
        if category == "vk11":
            params = {
                "start_date": start_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "end_date": end_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "limit": size,
                "offset": offset,
            }
            params_count = {
                "start_date": start_dt.strftime("%Y-%m-%d %H:%M:%S"),
                "end_date": end_dt.strftime("%Y-%m-%d %H:%M:%S"),
            }

        query_map = {
            "sellin": (QUERY_ERRO_SELLIN_PAGINATED, QUERY_ERRO_SELLIN),
            "sellin_detalhado": (
                QUERY_ERRO_SELLIN_DETALHADO,
                QUERY_ERRO_SELLIN_DETALHADO,
            ),
            "clientes": (QUERY_ERRO_CLIENTES_PAGINATED, QUERY_ERRO_CLIENTES),
            "produtos": (QUERY_ERRO_PRODUTOS_PAGINATED, QUERY_ERRO_PRODUTOS),
            "cutoff": (QUERY_ERRO_CUTOFF_PAGINATED, QUERY_ERRO_CUTOFF),
            "usuarios": (QUERY_ERRO_USUARIOS_PAGINATED, QUERY_ERRO_USUARIOS),
            "pagamentos": (
                QUERY_ERRO_PAGAMENTOS_LIST_PAGINATED,
                QUERY_ERRO_PAGAMENTOS_LIST,
            ),
            "pagamentos_sucesso": (
                QUERY_PAGAMENTOS_SUCESSO_LIST_PAGINATED,
                QUERY_PAGAMENTOS_SUCESSO_LIST,
            ),
            "vk11": (QUERY_ERRO_VK11_LIST_PAGINATED, QUERY_ERRO_VK11_LIST),
            "zaju": (QUERY_ERRO_ZAJU_LIST_PAGINATED, QUERY_ERRO_ZAJU_LIST),
        }

        if category not in query_map:
            raise HTTPException(status_code=400, detail="Categoria invalida")

        paginated_query_orig, count_query = query_map[category]

        if sort_by:
            import re

            # Validacao rigorosa contra SQL Injection
            if not re.match(r"^[a-zA-Z0-9_]+$", sort_by):
                raise HTTPException(
                    status_code=400, detail="Parametro de ordenacao invalido"
                )

            valid_order = "ASC" if order.upper() == "ASC" else "DESC"

            # Refatoracao Defensiva: Usamos uma subquery limpa.
            # O sort_by e validado via regex acima.
            wrapped_sql = f"SELECT * FROM ({count_query.text.strip().rstrip(';')}) AS sorted_subquery ORDER BY {sort_by} {valid_order} LIMIT :limit OFFSET :offset;"
            final_query = text(wrapped_sql)
        else:
            final_query = paginated_query_orig

        # Otimizacao: Usar COUNT(1) no banco ao inves de carregar tudo na memoria do Python
        # Refatoracao Segura:
        count_sql = f"SELECT COUNT(1) FROM ({count_query.text.strip().rstrip(';')}) AS total_count"
        count_res = await db.execute(text(count_sql), params_count)
        total_count = count_res.scalar() or 0

        # Obter os dados paginados
        res = await db.execute(final_query, params)
        rows = [dict(row) for row in res.mappings().all()]

        return {
            "source": "postgresql",
            "category": category,
            "data": rows,
            "page": page,
            "size": size,
            "total_count": total_count,
            "total_pages": (total_count + size - 1) // size,
        }
    except Exception as e:
        logger.error(f"Erro ao buscar detalhes da categoria {category}: {e}")
        # Mantemos o Mock fallback mas sem expor erro tecnico
        mock_data = []
        for i in range(offset, min(offset + size, 35)):  # Mock total of 35 items
            mock_data.append(
                {
                    "id": i,
                    "mensagem": f"Erro simulado {i} na integracao de {category}",
                    "data_emissao": "2026-03-17",
                }
            )

        return {
            "source": "mock",
            "warning": "Banco inacessivel",
            "category": category,
            "data": mock_data,
            "page": page,
            "size": size,
            "total_count": 35,
            "total_pages": 2,
        }


@router.get("/dashboard/details/vk11")
async def get_vk11_details(
    start_date: str = None,
    end_date: str = None,
    db: AsyncSession = Depends(get_db),
    user: str = Depends(get_current_user),
):
    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)
        params_str = {
            "start_date": start_dt.strftime("%Y-%m-%d %H:%M:%S"),
            "end_date": end_dt.strftime("%Y-%m-%d %H:%M:%S"),
        }

        result = await db.execute(QUERY_ORCAMENTO_INTEGRACAO, params_str)
        rows = [dict(row) for row in result.mappings().all()]

        return {"source": "postgresql", "data": rows}
    except Exception as e:
        logger.error(f"Erro ao buscar detalhes VK11: {e}")
        raise HTTPException(status_code=500, detail=str(e))


async def bg_generate_zaju_report(
    start_dt: datetime, end_dt: datetime, email: str, nome: str
):
    """Funcao de background para gerar o Excel e enviar por e-mail"""
    try:
        # Precisamos de uma nova sessao para o background task
        async with AsyncSessionLocal() as db:
            params = {"start_date": start_dt, "end_date": end_dt}
            result = await db.execute(QUERY_RELATORIO_ZAJU, params)
            rows = result.mappings().all()

            if not rows:
                # Opcional: enviar e-mail avisando que nao ha dados
                return

            # Garantir a ordenacao exata (keys do dict podem quebrar a ordem dependendo da versao do pandas)
            df = pd.DataFrame([dict(r) for r in rows])

            # Ajuste de pontuacao para o Padrao BRL (Virgula em decimais, ponto em milhares)
            colunas_numericas = [
                "Valor Bruto",
                "Valor Liquido",
                "Provisao Original",
                "Provisao % SAP",
                "Contrato COM % Bruto",
                "Contrato COM % Liquido",
                "Contrato LOG % Bruto",
                "Contrato LOG % Liquido",
                "Contrato CRE % Bruto",
                "Contrato CRE % Liquido",
                "Valor Provisao",
            ]
            for col in colunas_numericas:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").apply(
                        lambda x: (
                            f"{x:,.2f}".replace(",", "X")
                            .replace(".", ",")
                            .replace("X", ".")
                            if pd.notna(x)
                            else ""
                        )
                    )

            # Ajuste para campos puramente Percentuais
            colunas_percentuais = ["% Original", "% Atual"]
            for col in colunas_percentuais:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").apply(
                        lambda x: (
                            f"{x:,.2f}".replace(",", "X")
                            .replace(".", ",")
                            .replace("X", ".")
                            + "%"
                            if pd.notna(x)
                            else ""
                        )
                    )

            # Separar Contrato vs Promo & Acoes
            if not df.empty and "id_tipo_verba" in df.columns:
                df_contrato = df[df["id_tipo_verba"].isin([9, 10])].drop(
                    columns=["id_tipo_verba"]
                )
                df_promo = df[df["id_tipo_verba"].isin([5, 6])].drop(
                    columns=["id_tipo_verba"]
                )
            else:
                df_contrato = pd.DataFrame()
                df_promo = pd.DataFrame()

            ordem_colunas = [
                "Orcamento",
                "ID Ajuste Verba",
                "Linha de Investimento",
                "Tipo Linha Investimento",
                "Cod. Cliente",
                "Nome Cliente",
                "Nº Nota Fiscal",
                "VKORG",
                "Nº Documento",
                "Valor Bruto",
                "Valor Liquido",
                "% Original",
                "Provisao Original",
                "Provisao % SAP",
                "Contrato COM % Bruto",
                "Contrato COM % Liquido",
                "Contrato LOG % Bruto",
                "Contrato LOG % Liquido",
                "Contrato CRE % Bruto",
                "Contrato CRE % Liquido",
                "% Atual",
                "Valor Provisao",
                "Cutoff",
                "Data Criacao",
                "Purch. No C",
                "Tipo Doc",
                "Sequencial",
                "Cod. Material",
                "Material",
                "Unidade Medida",
                "Cond. Type",
                "Moeda",
                "Status",
                "Numfat Integracao",
                "Numov Integracao",
                "Data Integracao",
                "Mensagem Retorno",
            ]

            if not df_contrato.empty:
                df_contrato = df_contrato[
                    [c for c in ordem_colunas if c in df_contrato.columns]
                ]
            if not df_promo.empty:
                df_promo = df_promo[[c for c in ordem_colunas if c in df_promo.columns]]

            output = io.BytesIO()
            from openpyxl.styles import PatternFill, Font

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                if not df_contrato.empty:
                    df_contrato.to_excel(
                        writer, index=False, sheet_name="Verbas de Contrato"
                    )
                else:
                    pd.DataFrame(
                        columns=(
                            df.columns.drop("id_tipo_verba")
                            if "id_tipo_verba" in df.columns
                            else df.columns
                        )
                    ).to_excel(writer, index=False, sheet_name="Verbas de Contrato")

                if not df_promo.empty:
                    df_promo.to_excel(writer, index=False, sheet_name="Promo & Acoes")
                else:
                    pd.DataFrame(
                        columns=(
                            df.columns.drop("id_tipo_verba")
                            if "id_tipo_verba" in df.columns
                            else df.columns
                        )
                    ).to_excel(writer, index=False, sheet_name="Promo & Acoes")

                # Aplicar Identidade Visual Suzano em todas as abas
                for sheet_name in writer.sheets:
                    apply_excel_premium_style(writer, sheet_name)

            filename = f"relatorio_zaju_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"
            await mail_service.send_report_email(
                email_to=email,
                nome=nome,
                report_name="Relatorio Mensal ZAJU",
                file_content=output.getvalue(),
                filename=filename,
            )
            logger.info(f"Relatorio ZAJU enviado com sucesso para {email}")

    except Exception as e:
        logger.error(f"Erro no background task de exportacao ZAJU: {e}")


@router.get("/export/saldos")
@limiter.limit("10/minute")
async def export_saldos_report(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)

        params = {"start_date": start_dt, "end_date": end_dt}

        result = await db.execute(QUERY_RELATORIO_SALDOS, params)
        rows = result.mappings().all()

        if not rows:
            # Fallback para o usuario nao baixar arquivo vazio sem aviso
            df = pd.DataFrame(
                [{"Aviso": "Nenhum saldo encontrado para o periodo selecionado"}]
            )
        else:
            df = pd.DataFrame(rows)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet_name = "Saldos Disponiveis"
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            apply_excel_premium_style(writer, sheet_name)

        output.seek(0)
        filename = f"Saldos_Disponiveis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        logger.error(f"Erro ao exportar Saldos Disponiveis: {e}")
        raise HTTPException(status_code=500, detail="Erro ao gerar relatorio de saldos")


@router.get("/export/zaju")
@limiter.limit("10/minute")
async def export_zaju_report(
    request: Request,
    background_tasks: BackgroundTasks,
    start_date: str = None,
    end_date: str = None,
    user: dict = Depends(get_current_user),
):
    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)

        # Dispara o processamento em background
        background_tasks.add_task(
            bg_generate_zaju_report, start_dt, end_dt, user.email, user.nome
        )

        return {
            "status": "success",
            "message": f"O relatorio esta sendo gerado e sera enviado para {user.email} em instantes.",
        }

    except Exception as e:
        logger.error(f"Erro ao solicitar exportacao ZAJU: {e}")
        raise HTTPException(status_code=500, detail=str(e))


async def bg_generate_cg_elegiveis_report(
    start_dt: datetime,
    end_dt: datetime,
    email: str,
    nome: str,
    filter_month: str = None,
):
    try:
        from openpyxl.styles import PatternFill, Font

        async with AsyncSessionLocal() as db:
            params = {"start_date": start_dt, "end_date": end_dt}
            result = await db.execute(QUERY_RELATORIO_CG_ELEGIVEIS, params)
            rows = result.mappings().all()

            if not rows:
                return

            df = pd.DataFrame([dict(r) for r in rows])

            import io

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="CGs_Elegiveis")

                apply_excel_premium_style(writer, "CGs_Elegiveis")

            filename = f"relatorio_cg_elegiveis_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"

            # Construir a mensagem do periodo para o E-mail
            custom_period = None
            if filter_month:
                meses = [
                    "",
                    "Janeiro",
                    "Fevereiro",
                    "Marco",
                    "Abril",
                    "Maio",
                    "Junho",
                    "Julho",
                    "Agosto",
                    "Setembro",
                    "Outubro",
                    "Novembro",
                    "Dezembro",
                ]
                ano, mes = filter_month.split("-")
                mes_extenso = meses[int(mes)]
                str_start = f"{meses[start_dt.month]}/{start_dt.year}"
                str_end = f"{meses[end_dt.month]}/{end_dt.year}"
                if str_start != str_end:
                    elegibilidade = f"{str_start} a {str_end}"
                else:
                    elegibilidade = str_start
                custom_period = f"<strong>{mes_extenso}/{ano}</strong> (periodo de elegibilidade: {elegibilidade})"

            # Send email
            await mail_service.send_report_email(
                email_to=email,
                nome=nome,
                report_name="CGs Elegiveis Verba Fixa",
                file_content=output.getvalue(),
                filename=filename,
                custom_period_text=custom_period,
            )
            logger.info(f"Relatorio CGs Elegiveis enviado com sucesso para {email}")

    except Exception as e:
        logger.error(f"Erro no background task de exportacao CGs Elegiveis: {e}")


@router.get("/export/cg-elegiveis")
@limiter.limit("10/minute")
async def export_cg_elegiveis_report(
    request: Request,
    background_tasks: BackgroundTasks,
    start_date: str = None,
    end_date: str = None,
    filter_month: str = None,
    user: dict = Depends(get_current_user),
):
    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)

        background_tasks.add_task(
            bg_generate_cg_elegiveis_report,
            start_dt,
            end_dt,
            user.email,
            user.nome,
            filter_month,
        )

        return {
            "status": "success",
            "message": f"O relatorio esta sendo gerado e sera enviado para {user.email} em instantes.",
        }

    except Exception as e:
        logger.error(f"Erro ao solicitar exportacao CGs Elegiveis: {e}")
        raise HTTPException(
            status_code=500, detail="Erro ao processar solicitacao de exportacao."
        )


@router.get("/export/sellin-detalhado")
@limiter.limit("10/minute")
async def export_sellin_detailed(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    user: dict = Depends(get_current_user),
):
    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)

        async with AsyncSessionLocal() as db:
            params = {"start_date": start_dt, "end_date": end_dt}

            # --- Aba 1: Resumo (Dados do Modal) ---
            result_resumo = await db.execute(QUERY_ERRO_SELLIN, params)
            rows_resumo = result_resumo.mappings().all()

            if not rows_resumo:
                df_resumo = pd.DataFrame(
                    columns=[
                        "Erros",
                        "Data Emissao",
                        "Cliente",
                        "Nota Fiscal",
                        "Nº Documento",
                        "Tipo Doc Faturamento",
                    ]
                )
            else:
                raw_resumo_df = pd.DataFrame([dict(r) for r in rows_resumo])
                df_resumo = raw_resumo_df[
                    [
                        "erros",
                        "data_emissao",
                        "cliente",
                        "nro_nota_fiscal",
                        "nro_documento",
                        "tipo_doc_fat",
                    ]
                ].copy()

                df_resumo.columns = [
                    "Erros",
                    "Data Emissao",
                    "Cliente",
                    "Nota Fiscal",
                    "Nº Documento",
                    "Tipo Doc Faturamento",
                ]

            # --- Aba 2: Detalhamento (Dados Atuais) ---
            result_detalhado = await db.execute(QUERY_ERRO_SELLIN_DETALHADO, params)
            rows_detalhado = result_detalhado.mappings().all()

            if not rows_detalhado:
                df_detalhado = pd.DataFrame(
                    columns=[
                        "Erros",
                        "Data Registro",
                        "Data Emissao",
                        "Nota Fiscal",
                        "Nº Documento",
                        "Item",
                        "Cod. Cliente",
                        "Nome Cliente",
                        "Cliente",
                        "ID Produto",
                        "Nome Produto",
                        "Unidade",
                        "Quantidade",
                        "Valor Total",
                        "Valor Liquido",
                        "Vencimento",
                        "Tipo Doc",
                        "Referência Faturamento",
                    ]
                )
            else:
                raw_detalhado_df = pd.DataFrame([dict(r) for r in rows_detalhado])
                # Mapeamento para nomes amigaveis e colunas solicitadas
                df_detalhado = raw_detalhado_df[
                    [
                        "erros",
                        "dta_criacao",
                        "data_emissao",
                        "nro_nota_fiscal",
                        "nro_documento",
                        "item_documento",
                        "id_cliente",
                        "nom_cliente",
                        "cliente",
                        "id_produto",
                        "nom_produto",
                        "unidade_medida",
                        "quantidade",
                        "valor_total",
                        "valor_liquido",
                        "dta_venc_liq",
                        "tipo_doc_fat",
                        "referencia_fat",
                    ]
                ].copy()

                df_detalhado.columns = [
                    "Erros",
                    "Data Registro",
                    "Data Emissao",
                    "Nota Fiscal",
                    "Nº Documento",
                    "Item",
                    "Cod. Cliente",
                    "Nome Cliente",
                    "Cliente",
                    "ID Produto",
                    "Nome Produto",
                    "Unidade",
                    "Quantidade",
                    "Valor Total",
                    "Valor Liquido",
                    "Vencimento",
                    "Tipo Doc",
                    "Referência Faturamento",
                ]

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                # Primeira aba: Resumo Inconsistencias
                df_resumo.to_excel(
                    writer, index=False, sheet_name="Resumo Inconsistencias"
                )
                apply_excel_premium_style(writer, "Resumo Inconsistencias")

                # Segunda aba: Detalhamento Inconsistencias
                df_detalhado.to_excel(
                    writer, index=False, sheet_name="Detalhamento Inconsistencias"
                )
                apply_excel_premium_style(writer, "Detalhamento Inconsistencias")

            filename = f"sellin_detalhado_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"

            output.seek(0)
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )

    except Exception as e:
        logger.error(f"Erro ao exportar sellin detalhado: {e}")
        from fastapi.responses import JSONResponse

        return JSONResponse(
            status_code=500, content={"status": "error", "message": str(e)}
        )


@router.get("/export/clientes-detalhado")
@limiter.limit("10/minute")
async def export_clientes_detailed(
    request: Request,
    start_date: str = None,
    end_date: str = None,
    user: dict = Depends(get_current_user),
):
    try:
        start_dt, end_dt = parse_date_range(start_date, end_date)

        async with AsyncSessionLocal() as db:
            params = {"start_date": start_dt, "end_date": end_dt}

            # --- Dados do Relatorio ---
            result = await db.execute(QUERY_ERRO_CLIENTES, params)
            rows = result.mappings().all()

            if not rows:
                df = pd.DataFrame(
                    columns=[
                        "Erros",
                        "Data Registro",
                        "Cod. Cliente",
                        "Cliente",
                        "CNPJ",
                        "ativo_inativo",
                        "contato_cliente",
                        "email_contato_cliente",
                        "telefone_contato_cliente",
                        "sap_pagador",
                        "cod_customer_group",
                        "customer_group",
                        "cod_canal",
                        "canal",
                        "sub_canal",
                        "cod_regional",
                        "regional",
                    ]
                )
            else:
                raw_df = pd.DataFrame([dict(r) for r in rows])
                # Mapeamento e Ordem de Colunas
                df = raw_df[
                    [
                        "erros",
                        "dta_criacao",
                        "cod_cliente",
                        "nom_cliente",
                        "cnpj",
                        "ativo_inativo",
                        "contato_cliente",
                        "email_contato_cliente",
                        "telefone_contato_cliente",
                        "sap_pagador",
                        "cod_customer_group",
                        "customer_group",
                        "cod_canal",
                        "canal",
                        "sub_canal",
                        "cod_regional",
                        "regional",
                    ]
                ].copy()

                # Formatacao de campos booleanos para labels amigaveis
                df["ativo_inativo"] = (
                    df["ativo_inativo"]
                    .map({True: "Ativo", False: "Inativo"})
                    .fillna("Inativo")
                )
                df["sap_pagador"] = (
                    df["sap_pagador"].map({True: "Sim", False: "Nao"}).fillna("Nao")
                )

                df.columns = [
                    "Erros",
                    "Data Registro",
                    "Cod. Cliente",
                    "Cliente",
                    "CNPJ",
                    "Status",
                    "Contato",
                    "Email",
                    "Telefone",
                    "SAP Recebedor",
                    "Cod. Customer Group",
                    "Customer Group",
                    "Cod. Canal",
                    "Canal",
                    "Sub Canal",
                    "Cod. Regional",
                    "Regional",
                ]

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                # Aba Única: Detalhamento Inconsistencias
                df.to_excel(
                    writer, index=False, sheet_name="Detalhamento Inconsistencias"
                )
                apply_excel_premium_style(writer, "Detalhamento Inconsistencias")

            filename = f"clientes_detalhado_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"

            output.seek(0)
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )

    except Exception as e:
        logger.error(f"Erro ao exportar clientes detalhado: {e}")
        from fastapi.responses import JSONResponse

        return JSONResponse(
            status_code=500,
            content={
                "status": "error",
                "message": "Erro tecnico ao processar exportacao de clientes",
            },
        )


@router.post("/export/styled")
@limiter.limit("20/minute")
async def export_styled_json(
    request: Request, payload: Dict[str, Any], user: dict = Depends(get_current_user)
):
    """
    Recebe um JSON com:
    - 'title': String (nome do arquivo)
    - 'sheets': Lista de objetos { 'name': str, 'data': list }
    Retorna um arquivo Excel estilizado.
    """
    try:
        title = payload.get("title", "Export")
        sheets = payload.get("sheets", [])

        if not sheets:
            # Fallback para formato anterior de aba unica
            data = payload.get("data")
            if data:
                sheets = [{"name": title, "data": data}]
            else:
                raise HTTPException(
                    status_code=400, detail="Nenhum dado fornecido para exportacao."
                )

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for s in sheets:
                name = s.get("name", "Sheet")[:30]
                df = pd.DataFrame(s.get("data", []))
                df.to_excel(writer, index=False, sheet_name=name)
                apply_excel_premium_style(writer, name)

        output.seek(0)
        filename = f"{title.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        logger.error(f"Erro na exportacao generica: {e}")
        raise HTTPException(status_code=500, detail=str(e))
